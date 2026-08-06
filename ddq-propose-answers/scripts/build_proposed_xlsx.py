#!/usr/bin/env python3
"""Render a color-coded DDQ proposed-answers workbook from a proposal.json.

The JUDGMENT (matching each question to the Answer Bank, choosing tier /
confidence, writing the answer) is done by Claude and captured in proposal.json.
This script is the deterministic part: it scrubs internal handling instructions
out of the customer-facing text, applies the shading convention, and writes a
Summary + Proposed Answers workbook that a human can review quickly.

Usage:
    python3 build_proposed_xlsx.py proposal.json "Customer_DDQ_PROPOSED_YYYY-MM-DD.xlsx"

proposal.json schema:
{
  "meta": {
    "customer": "...", "assessment": "...", "answer_bank": "...",
    "drafted": "YYYY-MM-DD", "notes": ["free-form bullets for the Summary sheet"]
  },
  "questions": [
    {
      "section":    "1. Corporate Information",
      "id":         "1.1",
      "parent":     "",                 # the id this hangs off, "" if top-level
      "question":   "Please provide ...",
      "type":       "radio" | "free-text",
      "options":    [{"label":"Yes","checked":true}, ...],   # [] for free-text
      "answer":     "Yes",              # the radio pick, or "" for free-text
      "rationale":  "The customer-facing answer text (verbatim from the bank, or synthesized).",
      "tier":       "1" | "2-3" | "synth",
      "confidence": 88,                 # int 0-100, REQUIRED for tier "synth", ignored otherwise
      "source":     "1. Commercial — \"Does the organization carry insurance...\"",
      "note":       "short internal action note (why gray, what to verify, gap, etc.)",
      "attachment": "<your-org> Certificate of Insurance 2026.pdf"   # or ""
    }
  ]
}

Shading (applied to the Proposed Answer + Rationale cells):
  tier "1"   -> no fill        (primary verbatim match; ready after a glance)
  tier "2-3" -> LIGHT YELLOW   (secondary source; verify it fits)
  tier "synth" -> by confidence: >90 no fill, 80-90 LIGHT GRAY, <80 LIGHT PINK

Copyright (C) 2026 Deborah Beckett

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
import json, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- scrub: internal handling instructions must never reach the customer ---
# These live in the Answer Bank comment fields on purpose (internal guidance);
# strip them from the customer-facing rationale, and report what was removed so
# nothing is cut silently. Keep legitimate answer language like "assessed on a
# case-by-case basis" — the tell is a routing/handling instruction aimed at us.
SCRUB_MARKERS = [
    "please escalate", "escalate to legal", "escalate to finance", "escalate to grc",
    "escalate to sales", "escalate to support",
    "handles approval", "approves case-by-case", "approved internally",
    "internal approval", "may be superseded", "trust package", "do not publish",
    "do not cite", "hold —", "blocked —", "needs input",
]
# Inline parentheticals that carry a routing/handling instruction (e.g.
# "(escalate to Legal)", "(approved internally, case-by-case)") — name-agnostic:
# match any (...) containing a handling verb, so notes get stripped regardless of
# who is named. Legitimate answer language like "(assessed on a case-by-case basis)"
# has no handling verb and is kept.
INLINE_HANDLING = re.compile(
    r"\s*\([^)]*\b(?:escalat|approv|handle[sd]?|internal|do not (?:publish|cite)"
    r"|superseded|hold|blocked|needs input)[^)]*\)", re.I)
def scrub(text):
    if not text:
        return text, []
    t = INLINE_HANDLING.sub("", text)
    parts = re.split(r"(?<=[.!?])\s+", t)
    kept, removed = [], []
    for p in parts:
        if any(m in p.lower() for m in SCRUB_MARKERS):
            removed.append(p.strip())
        else:
            kept.append(p)
    if t != text and not removed:
        removed.append("(inline internal-handling note in parentheses)")
    return " ".join(kept).strip(), removed

def band_fill(q, fills):
    tier = q.get("tier", "synth")
    if tier == "1":
        return None
    if tier == "2-3":
        return fills["yellow"]
    c = int(q.get("confidence", 0))
    if c > 90:
        return None
    if c >= 80:
        return fills["gray"]
    return fills["pink"]

def main(proposal_path, out_path):
    data = json.load(open(proposal_path))
    meta = data.get("meta", {})
    qs = data["questions"]

    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    fills = {
        "yellow": PatternFill("solid", fgColor="FFF3B0"),
        "gray":   PatternFill("solid", fgColor="EDEDED"),
        "pink":   PatternFill("solid", fgColor="FBE0E0"),
        "attach": PatternFill("solid", fgColor="E8F0FE"),
    }

    n_t1 = sum(1 for q in qs if q.get("tier") == "1")
    n_t23 = sum(1 for q in qs if q.get("tier") == "2-3")
    synth = [q for q in qs if q.get("tier") == "synth"]
    n_white = n_t1 + sum(1 for q in synth if int(q.get("confidence", 0)) > 90)
    n_gray = sum(1 for q in synth if 80 <= int(q.get("confidence", 0)) <= 90)
    n_pink = sum(1 for q in synth if int(q.get("confidence", 0)) < 80)
    n_att = sum(1 for q in qs if q.get("attachment"))

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    rows = [
        (f"{meta.get('customer','')} — Proposed DDQ Answers", ""),
        ("Assessment", meta.get("assessment", "")),
        ("Answer Bank", meta.get("answer_bank", "")),
        ("Drafted", meta.get("drafted", "")),
        ("", ""),
        ("Total questions", str(len(qs))),
        ("Tier-1 primary (verbatim, no shading)", str(n_t1)),
        ("Tier-2/3 secondary (light yellow)", str(n_t23)),
        ("Synthesized last-resort", str(len(synth))),
        ("", ""),
        ("SHADING KEY",
         "No shading = Tier-1 primary verbatim match (Answer Bank tabs starting '1'). "
         "LIGHT YELLOW = Tier-2/3 secondary source (tabs starting '2'/'3'). "
         "For synthesized answers: no shading = >90% confidence, LIGHT GRAY = 80-90%, LIGHT PINK = <80%. "
         "Review effort goes to the gray and pink rows first."),
        ("Confidence tally", f"Ready/verbatim (white): {n_white}   |   Verify (yellow+gray): {n_t23 + n_gray}   |   Judgment (pink): {n_pink}"),
        ("Attachments proposed", str(n_att)),
    ]
    for note in meta.get("notes", []):
        rows.append(("NOTE", note))
    for i, (a, b) in enumerate(rows, 1):
        ca = ws.cell(i, 1, a); cb = ws.cell(i, 2, b)
        ca.font = Font(bold=True, size=(13 if i == 1 else 10)); ca.alignment = wrap
        cb.alignment = wrap; cb.font = Font(size=10)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 112

    # ---- Proposed Answers sheet ----
    pa = wb.create_sheet("Proposed Answers")
    heads = ["Section", "Q #", "Hangs off", "Question", "Type", "Response Options",
             "Proposed Answer", "Proposed Comment / Rationale", "Source", "Confidence / Action", "Proposed Attachment"]
    for j, h in enumerate(heads, 1):
        c = pa.cell(1, j, h); c.fill = hdr_fill; c.font = hdr_font; c.alignment = wrap; c.border = border
    for j, w in enumerate([22, 6, 8, 46, 9, 22, 30, 60, 26, 34, 32], 1):
        pa.column_dimensions[chr(64 + j)].width = w
    pa.freeze_panes = "A2"

    scrub_log = []
    r = 2
    for q in qs:
        opts = q.get("options", [])
        is_radio = bool(opts) or q.get("type") == "radio"
        pick = q.get("answer", "")
        # An option is selected if its own `checked` flag is set (multi-select
        # checkbox) OR it matches the single `answer` pick (radio / Yes-No).
        def is_sel(o):
            return bool(o.get("checked")) or (pick and o["label"] == pick)
        if is_radio:
            optstr = "\n".join(f"[{'X' if is_sel(o) else ' '}] {o['label']}" for o in opts)
        else:
            optstr = "(free-text)"
        checked_labels = [o["label"] for o in opts if o.get("checked")]
        rationale, removed = scrub(q.get("rationale", ""))
        if removed:
            scrub_log.append((q["id"], removed))
        # Proposed Answer: single pick, else the joined multi-select picks, else
        # a clear "no selection" flag for a radio the reviewer must still settle.
        if pick:
            prop = pick
        elif checked_labels:
            prop = "; ".join(checked_labels)
        elif is_radio:
            prop = "(no selection — see rationale)"
        else:
            prop = "(free-text — see rationale)"
        tier = q.get("tier", "synth")
        if tier == "1":
            tiertxt = "Tier-1 (primary)"
        elif tier == "2-3":
            tiertxt = "Tier-2/3 (secondary)"
        else:
            tiertxt = f"Synthesized {q.get('confidence','?')}%"
        if q.get("note"):
            tiertxt += f" — {q['note']}"
        vals = [q.get("section", ""), q["id"], q.get("parent", ""), q["question"],
                "radio" if is_radio else "free-text", optstr, prop, rationale,
                q.get("source", ""), tiertxt, q.get("attachment", "")]
        for j, v in enumerate(vals, 1):
            c = pa.cell(r, j, v); c.alignment = wrap; c.border = border; c.font = Font(size=9)
        fill = band_fill(q, fills)
        if fill:
            pa.cell(r, 7).fill = fill; pa.cell(r, 8).fill = fill
        if q.get("attachment"):
            pa.cell(r, 11).fill = fills["attach"]
        r += 1

    wb.save(out_path)
    print(f"wrote {out_path}")
    print(f"questions: {len(qs)} | tier-1: {n_t1} | tier-2/3: {n_t23} | synth: {len(synth)} "
          f"(white {sum(1 for q in synth if int(q.get('confidence',0))>90)}, gray {n_gray}, pink {n_pink}) "
          f"| attachments: {n_att}")
    if scrub_log:
        print(f"\nSCRUBBED internal-only text from {len(scrub_log)} customer-facing answer(s):")
        for qid, rem in scrub_log:
            for frag in rem:
                print(f"   {qid}: removed -> {frag[:110]}")
    else:
        print("no internal-instruction text needed scrubbing")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: build_proposed_xlsx.py <proposal.json> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
