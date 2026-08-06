#!/usr/bin/env python3
"""
Build a portal-questionnaire extract workbook from a JSON capture.

Usage:
    python3 build_extract_xlsx.py capture.json "/path/to/output.xlsx"

The JSON shape (see references/capture-schema.md for the full spec):

{
  "meta": {
    "customer": "<Customer>",
    "assessment": "Vendor Service Risk Assessment v1.3",
    "record": "<record-id>",
    "platform": "Fusion Framework (FFQ) on Salesforce Experience Cloud",
    "portal_url": "<portal URL>",
    "respondent": "<name of the person who completed it>",
    "completion": "100% (Business Resilience 7/7, Disaster Recovery 16/16)",
    "extracted": "2026-07-14",
    "notes": ["free-form strings, each becomes a bullet on the Summary sheet"]
  },
  "questions": [
    {
      "section": "Disaster Recovery",
      "id": "DR-14",                 # portal number OR stable control id
      "parent": "DR-1",              # id this question hangs off, "" if top-level
      "question": "Are your production and backup environments ...",
      "answer": "Yes",               # selected option label, or free-text value
      "options": [                   # every option the portal offered
        {"label": "Yes", "checked": true},
        {"label": "No",  "checked": false}
      ],
      "comment": "Provided through our hosting provider ...",  # comment / rationale, "" if none
      "instruction": "",             # portal's "please upload X" text, "" if none
      "attachment": ""               # what was attached + what it appears to be, "" if none
    }
  ]
}

Design choices worth keeping:
- The "Response Options Available" column lists EVERY option with [X] on the selected
  one. This is what lets someone later spot where a "least-worst-fit" answer was picked.
- Free-text questions (no radio/checkbox) get options == [] and render as
  "(free-text response - no options)".
- The Attachment column always renders explicitly ("- none -" when empty) so a blank
  never reads as "not checked".

Copyright (C) 2026 Deborah Beckett

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FREE_TEXT = "(free-text response - no options)"


def build(capture_path, out_path):
    with open(capture_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    meta = data.get("meta", {})
    questions = data.get("questions", [])

    wb = Workbook()

    # ---------------- Summary sheet ----------------
    ws = wb.active
    ws.title = "Summary"
    title_font = Font(bold=True, size=14)
    hdr_font = Font(bold=True)

    n_attach = sum(1 for q in questions if (q.get("attachment") or "").strip())
    n_instr = sum(1 for q in questions if (q.get("instruction") or "").strip())
    sections = []
    for q in questions:
        s = q.get("section", "")
        if s and s not in sections:
            sections.append(s)

    title = meta.get("customer", "Portal Questionnaire") + " - Extract"
    summary = [
        (title, ""),
        ("Extracted", meta.get("extracted", "")),
        ("", ""),
        ("Customer", meta.get("customer", "")),
        ("Assessment", meta.get("assessment", "")),
        ("Portal record", meta.get("record", "")),
        ("Portal platform", meta.get("platform", "")),
        ("Portal URL", meta.get("portal_url", "")),
        ("Respondent account", meta.get("respondent", "")),
        ("Completion at extraction", meta.get("completion", "")),
        ("Sections", ", ".join(sections)),
        ("Total questions", str(len(questions))),
        ("Questions requesting a file upload", str(n_instr)),
        ("Files actually attached", str(n_attach) if n_attach else "None"),
        ("", ""),
    ]
    for label in meta.get("notes", []):
        summary.append(("Note", label))

    for r, (a, b) in enumerate(summary, start=1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        if r == 1:
            ca.font = title_font
        elif a and b and a != "Note":
            ca.font = hdr_font
        elif a == "Note":
            ca.font = hdr_font
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 95

    # ---------------- Q&A sheet ----------------
    qa = wb.create_sheet("Q&A")
    headers = ["Section", "Q #", "Hangs off", "Question", "Our Answer",
               "Response Options Available", "Comment / Rationale",
               "Instruction (portal ask)", "Attachment / File"]
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    for c, h in enumerate(headers, start=1):
        cell = qa.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border

    # alternate a faint fill per section so the eye can group them
    palette = ["E8EEF7", "FBEEE6", "EAF3EA", "F5EEF6", "FDF6E3", "EAF0F2"]
    sec_fill = {s: palette[i % len(palette)] for i, s in enumerate(sections)}

    for i, q in enumerate(questions, start=2):
        options = q.get("options") or []
        answer = q.get("answer", "")
        if options:
            opts_str = "\n".join(
                ("[X] " if o.get("checked") else "[  ] ") + o.get("label", "")
                for o in options
            )
        else:
            opts_str = FREE_TEXT
        attach = (q.get("attachment") or "").strip()
        vals = [
            q.get("section", ""),
            q.get("id", ""),
            q.get("parent", ""),
            q.get("question", ""),
            answer,
            opts_str,
            q.get("comment", ""),
            q.get("instruction", ""),
            attach if attach else "- none -",
        ]
        for c, v in enumerate(vals, start=1):
            cell = qa.cell(row=i, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if c == 1:
                cell.fill = PatternFill("solid", fgColor=sec_fill.get(q.get("section", ""), "FFFFFF"))
        # italicize the answer when it's an N/A or a free-text response, for quick scanning
        if answer.startswith("N/A") or not options:
            qa.cell(row=i, column=5).font = Font(italic=True)
        # flag rows that have an outstanding upload ask but nothing attached
        if (q.get("instruction") or "").strip() and not attach:
            qa.cell(row=i, column=9).font = Font(color="C00000")

    widths = [16, 8, 11, 55, 30, 46, 40, 34, 20]
    for c, w in enumerate(widths, start=1):
        qa.column_dimensions[get_column_letter(c)].width = w
    qa.freeze_panes = "A2"
    qa.row_dimensions[1].height = 30

    wb.save(out_path)
    print(f"wrote {out_path}")
    print(f"questions: {len(questions)} | upload asks: {n_instr} | files attached: {n_attach}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: build_extract_xlsx.py <capture.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])
